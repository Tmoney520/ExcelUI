# -*- coding: utf-8 -*- 
"""          
# @Time : 2023/3/21 21:24
#  :TGW
"""
import openpyxl
from web_keys.webkey import Keys
from common.logger import test_log

#处理测试用例中参数单元格的函数，处理成json格式     或者直接规定只能在excel中输入json格式
def arguments(value):
        data = {}
        #如果value有值，进行切分
        if value:
                str_temp = value.split(';')
                for temp in str_temp:
                        t = temp.split('=', 1)
                        data[t[0]] = t[1]
        #如果value没值，不做任何操作
        else:
                pass
        return data

def read(log,file):

        #加载excel文件
        excel = openpyxl.load_workbook(file)

        # 获取所有的sheet页，来执行里面的测试内容
        for name in excel.sheetnames:
                log.info('**********正在执行{}流程*********'.format(name))
                sheet = excel[name]
                try:
                        #获取到每个sheet页的值
                        for values in sheet.values:

                                # 获取测试用例的正文内容
                                if type(values[0]) is int:

                                        # 用例描述可以用于日志的输出
                                        log.info('*****************正在执行：{}*****************'.format(values[3]))

                                        # 参数的处理:通过一个dict来接收所有的参数内容。便于定值不定长的传参形态
                                        data = arguments(values[2])

                                        # 实例化操作
                                        if values[1] == 'open_browser':
                                                key = Keys(**data)

                                        # 断言行为:基于断言的返回结果来判定测试的成功失败，并进行写入操作
                                        elif 'assert' in values[1]:
                                                status = getattr(key, values[1])(text=values[4], **data)
                                                # 基于status判定写入的测试结果
                                                if status:
                                                        # excel_conf.pass_(sheet.cell, row=values[0] + 2, column=6)
                                                        sheet.cell(row=values[0] + 2, column=6).value = 'Pass'
                                                else:
                                                        # excel_conf.failed(sheet.cell, row=values[0] + 2, column=6)
                                                        sheet.cell(row=values[0] + 2, column=6).value = 'Fail'

                                                # 保存Excel：放在这里以确保每一次写入都可以被保存，避免因为代码报错而未保存之前的测试结果
                                                excel.save(file)

                                        # 常规操作行为
                                        else:
                                                getattr(key, values[1])(**data)
                except:
                        continue

                log.info('**********{}流程执行结束*********'.format(name))
        excel.close()
